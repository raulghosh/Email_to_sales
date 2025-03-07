from openpyxl.styles import Alignment, PatternFill, Font, numbers
from openpyxl.utils import get_column_letter
from typing import List
import pandas as pd
import numpy as np

def format_excel_sheet(worksheet, df: pd.DataFrame,  sheet_name: str, sales_rep=True) -> None:
    """Apply formatting to an Excel worksheet."""
    # Style the header row
    header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    
    # Right-align numerical columns
    numerical_columns = [
        col for col in df.columns 
        if any(keyword in col.lower() for keyword in ['sales', 'opp', 'margin'])
    ]
    
    for col_name in numerical_columns:
        col_idx = df.columns.get_loc(col_name)
        # Ensure col_idx is an integer
        
        if isinstance(col_idx, (np.ndarray, list)):
            col_idx = int(col_idx[0])
        else:
            col_idx = int(col_idx)
            
        # Apply number formatting
        for row in worksheet.iter_rows(min_row=2, min_col=int(col_idx+1), max_col=int(col_idx+1)):
            for cell in row:
                cell.alignment = Alignment(horizontal="right")
                if "margin" in col_name.lower():
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                else:
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Format "Last Trans. Date" column
    if "Last Trans. Date" in df.columns:
        date_col_idx = df.columns.get_loc("Last Trans. Date") + 1
        for row in worksheet.iter_rows(min_row=2, min_col=date_col_idx, max_col=date_col_idx):
            for cell in row:
                cell.number_format = "MM/DD/YYYY"

    # Adjust column widths
    for col_num, column in enumerate(df.columns, 1):
        max_length = max(df[column].astype(str).apply(len).max(), len(column)) + 2
        worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_length, 30)
        
        if "Item Desc" in df.columns:
            item_desc_col_idx = df.columns.get_loc("Item Desc") + 1
            max_length = max(df["Item Desc"].astype(str).apply(len).max(), len("Item Desc")) + 12
            worksheet.column_dimensions[get_column_letter(item_desc_col_idx)].width = min(max_length, 50)
        worksheet.auto_filter.ref = worksheet.dimensions
    if sales_rep:
        worksheet.freeze_panes = "G2"
    elif sheet_name not in ["Attic", "Basement"]:
        worksheet.freeze_panes = None
    else:
        worksheet.freeze_panes = "H2"