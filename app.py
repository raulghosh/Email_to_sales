import os
import datetime as dt
import pandas as pd
from dotenv import load_dotenv
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl.styles import Alignment,PatternFill, Font
from openpyxl.utils import get_column_letter

# Load environment variables from .env file
load_dotenv()

# Configuration
main_folder = os.getenv("MAIN_FOLDER")  # Folder containing the exported file
input_file_name = "Sales_Report_Temp.xlsx"  # exported file name
output_folder = os.path.join(main_folder, "Filtered_Reports")  # Folder to save filtered files
input_file = os.path.join(main_folder, input_file_name)
email_config = {
    "smtp_server": "relay.int.distco.com",
    "smtp_port": 25,  # Port 25 is commonly used for relay servers
    "sender_email": os.getenv("EMAIL_USER"),
    "test_email": "rghosh@veritivcorp.com"
}

print(f"Email User: {email_config['sender_email']}")
# Get the current month and year
current_date = dt.datetime.now()
month_year = current_date.strftime("%b, %Y")

# Load the exported file
data = pd.read_excel(input_file)
print(f"Data loaded from {input_file}")

# Filter out rows where Sales Rep Email or Manager Email is NaN
data = data.dropna(subset=["Rep Email", "Manager Email"])


# Format columns
sales_columns = [col for col in data.columns if 'sales' in col.lower()]
for col in sales_columns:
    data[col] = data[col].fillna(0).round(0).astype(int)  # Ensure no decimals

opp_columns = [col for col in data.columns if 'opp' in col.lower()]
for col in opp_columns:
    data[col] = data[col].fillna(0).round(0).astype(int)  # Ensure no decimals

margin_columns = [col for col in data.columns if 'margin' in col.lower()]
for col in margin_columns:
    data[col] = (data[col].fillna(0) * 100).round(1)

# Apply formatting for display/export
formatted_data = data.copy()
for col in sales_columns + opp_columns:
    # Use comma-separated format without trailing zeros
    formatted_data[col] = formatted_data[col].apply(lambda x: f"{x:,}") 

for col in margin_columns:
    formatted_data[col] = formatted_data[col].apply(lambda x: f"{x:.1f}%")
    
print(f"Data formatted for display")
    
# Get unique sales rep emails and names, limited to the first 3
sales_reps = data[["Rep Email", "Rep Name"]].drop_duplicates().head(1).set_index("Rep Email")["Rep Name"].to_dict()

# Get unique managers emails and names, limited to the first 3
managers = data[["Manager Email", "Manager Name"]].drop_duplicates().head(1).set_index("Manager Email")["Manager Name"].to_dict()

# Create output folder if it doesn't exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Save the formatted data to an Excel file with filters on the header
output_file = os.path.join(output_folder, f"Formatted_Sales_Report_{month_year}.xlsx")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    formatted_data.to_excel(writer, index=False, sheet_name='Sales Report')
    worksheet = writer.sheets['Sales Report']
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max(max_length + 2, len(col[0].value)), 30)
        worksheet.column_dimensions[column].width = adjusted_width
    worksheet.auto_filter.ref = worksheet.dimensions

# Function to send email
def send_email(to_email, subject, body, attachment_path=None):
    msg = MIMEMultipart()
    msg["From"] = email_config["sender_email"]
    msg["To"] = to_email
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "html"))

    if attachment_path:
        with open(attachment_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(attachment_path)}")
            msg.attach(part)

    with smtplib.SMTP(email_config["smtp_server"], email_config["smtp_port"]) as server:
        server.sendmail(email_config["sender_email"], to_email, msg.as_string())

print("Email function defined")

# Process and send emails to sales reps
for email, name in sales_reps.items():
    # Use the ORIGINAL DATA (not formatted_data) for calculations
    filtered_data_raw = data[data["Rep Email"] == email]  # <-- Use raw data for sums
    filtered_data_formatted = formatted_data[formatted_data["Rep Email"] == email]  # <-- For saving to Excel

    # Remove columns from formatted data
    filtered_data_formatted = filtered_data_formatted.drop(columns=["Rep Email", "Rep Name", "Manager Email", "Manager Name"])

    cols = list(filtered_data_formatted.columns)
    ltm_index = cols.index('LTM Gross Sales')
    opp_index = cols.index('Opp to Floor')
    cols.insert(ltm_index + 1, cols.pop(opp_index))
    filtered_data_formatted = filtered_data_formatted[cols]

    # Save the formatted data to a new file WITH RIGHT-ALIGNED NUMBERS
    output_file = os.path.join(output_folder, f"{name}_Report.xlsx")
    
    # Apply filter to the header row
    worksheet.auto_filter.ref = worksheet.dimensions
    
    # Color the header row in dark green with white font
    header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    left_alignment = Alignment(horizontal='left')
    right_alignment = Alignment(horizontal='right')
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = left_alignment  # Left-align headers
    
    # Use OpenPyXL to format cells
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        filtered_data_formatted.to_excel(writer, index=False, sheet_name='Sales Report')
        worksheet = writer.sheets['Sales Report']
        
        # Identify numerical columns (sales, opp, margin)
        numerical_columns = [
            col for col in filtered_data_formatted.columns 
            if any(keyword in col.lower() for keyword in ['sales', 'opp', 'margin'])
        ]
        
        # Format numerical columns
        for col_name in numerical_columns:
            col_idx = filtered_data_formatted.columns.get_loc(col_name)
            
            # Right-align data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx+1, max_col=col_idx+1):
                for cell in row:
                    cell.alignment = right_alignment

        # Set column widths
        for col_idx, column in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(col_idx + 1)
            
            # Include header in width calculation
            header_length = len(str(column[0].value))
            cell_lengths = [len(str(cell.value)) for cell in column[1:]]  # Exclude header
            max_cell_length = max(cell_lengths) if cell_lengths else 0
            
            # Use whichever is larger: header length or max data length
            calculated_width = max(header_length, max_cell_length)
            adjusted_width = min(calculated_width + 2, 30)  # Max 30 characters
            
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Apply filter to header row
        worksheet.auto_filter.ref = worksheet.dimensions
        
    # Calculate sums using RAW DATA
    basement_count = filtered_data_raw[filtered_data_raw["Region"] == "Basement"].shape[0]
    attic_count = filtered_data_raw[filtered_data_raw["Region"] == "Attic"].shape[0]
    KVI_count = filtered_data_raw[(filtered_data_raw["KVI Type"] == "2: KVI") | (filtered_data_raw["KVI Type"] == "3: Super KVI")].shape[0]
    basement_sales = filtered_data_raw[filtered_data_raw["Region"] == "Basement"]["LTM Gross Sales"].sum()
    attic_sales = filtered_data_raw[filtered_data_raw["Region"] == "Attic"]["LTM Gross Sales"].sum()
    opp_to_floor = filtered_data_raw["Opp to Floor"].sum()

    # Format numbers using raw numeric values
    basement_sales_formatted = f"{basement_sales:,.0f}"
    attic_sales_formatted = f"{attic_sales:,.0f}"
    opp_to_floor_formatted = f"{opp_to_floor:,.0f}"

    # Format numbers with comma separation and round to zero decimals
    basement_sales_formatted = f"{basement_sales:,.0f}"
    attic_sales_formatted = f"{attic_sales:,.0f}"
    opp_to_floor_formatted = f"{opp_to_floor:,.0f}"

    # Create summary table by Region using RAW DATA (numeric)
    summary_table = filtered_data_raw.groupby("Region").agg({
        "LTM Gross Sales": "sum",
        "Opp to Floor": "sum"
    }).reset_index()

    # Format the summary table numbers (using the raw numeric values)
    summary_table["LTM Gross Sales"] = summary_table["LTM Gross Sales"].apply(lambda x: f"{x:,.0f}")
    summary_table["Opp to Floor"] = summary_table["Opp to Floor"].apply(lambda x: f"{x:,.0f}")

    # Convert summary table to HTML with right-aligned numbers
    summary_html = summary_table.to_html(index=False, classes='summary-table')
    
    # Add CSS for right-aligning numbers in the table
    summary_html = f"""
    <style>
        .summary-table th, .summary-table td {{
            text-align: right;
        }}
    </style>
    {summary_html}
    """

    body = f"""
    <div style="text-align: left;">
        <p>Hi {name},</p>
        <p>Hope you are doing good. Attached is the Attic and Basement Report for {month_year}.</p>
        <p>This month you have {basement_count} action items in 'Basement' corresponding to ${basement_sales_formatted} of gross sales and you have {attic_count} action items in 'Attic' corresponding to {attic_sales_formatted} of gross sales.</p>
        <p>Raising the items in basement to the recommended margin will result in ${opp_to_floor_formatted} of commission profit gain.</p>
        <p>Summary by Region:</p>
        {summary_html}
        <p>Thanks,<br>Pricing Team</p>
    </div>
    """

    # Send email with attachment
    send_email(
        to_email=email_config["test_email"],  # test email
        subject=f"{name}: Attic and Basement Report {month_year}",
        body=body,
        attachment_path=output_file,
    )
    print(f"Email sent to manager {name} at {email_config['test_email']}")